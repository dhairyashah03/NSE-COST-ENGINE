"""
Validation Report Generator
============================

Generates a comprehensive Excel report with:
  Sheet 1: Contract Note Validation (engine vs actual)
  Sheet 2: Per-Trade Breakdown
  Sheet 3: Stress Tests
  Sheet 4: Market Impact — Order Size Scenarios
  Sheet 5: Market Impact — Cross-Stock Comparison
  Sheet 6: Rate Schedule

Pulls live volatility and volume data from Dhan API for market impact.

Usage:
    python reports/generate_validation.py
    python reports/generate_validation.py --input tests/fixtures/contract_note_20260508.json
    python reports/generate_validation.py --output reports/my_report.xlsx
    python reports/generate_validation.py --broker zerodha
"""

import argparse
import json
import sys
import os
import math
from pathlib import Path
from datetime import datetime

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from nse_cost_engine import (
    CostEngine, Trade, Segment, TradeType,
    Side as TradeSide, Exchange,
)

# ---------------------------------------------------------------------------
# Formatting constants
# ---------------------------------------------------------------------------
FONT_NAME = "Calibri Light"
FONT_SIZE = 12
SC = 2  # start column (B)
SR = 2  # start row (2)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _check_openpyxl():
    try:
        import openpyxl
        return True
    except ImportError:
        print("openpyxl not installed. Install with: pip install openpyxl")
        return False


def build_trades_from_json(trade_list):
    trades = []
    for t in trade_list:
        kwargs = {
            "symbol": t["symbol"],
            "segment": Segment(t["segment"]),
            "trade_type": TradeType(t["trade_type"]),
            "side": TradeSide(t["side"]),
            "exchange": Exchange(t["exchange"]),
            "price": t["price"],
            "quantity": t["quantity"],
            "lot_size": t.get("lot_size", 1),
        }
        for key in ("premium", "strike_price", "settlement_price",
                     "mtf_funding_pct", "mtf_holding_days",
                     "daily_volatility", "avg_daily_volume"):
            if key in t:
                kwargs[key] = t[key]
        if t.get("is_exercise"):
            kwargs["is_exercise"] = True
        trades.append((t, Trade(**kwargs)))
    return trades


def run_validation(engine, trades, actuals):
    results = []
    for raw, trade in trades:
        result = engine.calculate(trade)
        results.append((raw, trade, result))

    combined = {}
    for _, trade, result in results:
        cn = result.as_contract_note_dict()
        for key, val in cn.items():
            combined[key] = combined.get(key, 0) + val

    mapping = [
        ("Brokerage", "brokerage", "brokerage"),
        ("NSE Transaction Charges", "exchange_charges", "nse_transaction_charges"),
        ("SEBI Fees", "sebi_fee", "sebi_fees"),
        ("GST (IGST 18%)", "gst", "gst_igst_18"),
        ("Stamp Duty", "stamp_duty", "stamp_duty"),
        ("STT", "stt", "stt"),
    ]

    comparisons = []
    for label, eng_key, cn_key in mapping:
        eng_val = combined.get(eng_key, 0)
        cn_val = actuals.get(cn_key, 0) if actuals else None
        delta = (eng_val - cn_val) if cn_val is not None else None
        match = abs(delta) < 0.10 if delta is not None else None
        comparisons.append({"label": label, "engine": eng_val,
                           "contract_note": cn_val, "delta": delta, "match": match})

    return results, comparisons


def print_console_report(data, results, comparisons):
    print("=" * 65)
    print(f"Contract Note Validation — {data.get('broker', '?')} | {data.get('contract_date', '?')}")
    print("=" * 65)
    print(f"\n{'Component':30s} {'Engine':>10s} {'CN':>10s} {'Match':>6s}")
    print("-" * 60)
    for c in comparisons:
        cn_str = f"₹{c['contract_note']:>8.2f}" if c['contract_note'] is not None else "    N/A"
        match_str = "✓" if c['match'] else "✗" if c['match'] is not None else " "
        print(f"  {c['label']:28s} ₹{c['engine']:>8.2f}  {cn_str}  {match_str}")
    total_eng = sum(c['engine'] for c in comparisons)
    total_cn = sum(c['contract_note'] for c in comparisons if c['contract_note'] is not None)
    print("-" * 60)
    print(f"  {'TOTAL':28s} ₹{total_eng:>8.2f}  ₹{total_cn:>8.2f}")
    print()


def fetch_live_impact_data(symbols_and_ids):
    """
    Pull live volatility and volume from Dhan API.
    Returns dict: {symbol: (daily_vol, avg_volume)} or None if unavailable.
    """
    try:
        from nse_cost_engine.data_feed import DhanDataFeed
        feed = DhanDataFeed()
        if not feed.is_configured:
            print("  ⚠ Dhan API not configured — using mock data for market impact")
            return None

        data = {}
        for symbol, sec_id in symbols_and_ids.items():
            try:
                vol, adv = feed.get_impact_params(sec_id)
                data[symbol] = (vol, adv)
                print(f"  ✓ {symbol}: vol={vol:.4f}, avg_vol={adv:,}")
            except Exception as e:
                print(f"  ✗ {symbol}: {e}")
        return data if data else None
    except ImportError:
        print("  ⚠ requests/python-dotenv not installed — using mock data")
        return None


# ---------------------------------------------------------------------------
# Excel Generation
# ---------------------------------------------------------------------------

def generate_excel_report(data, trades, results, comparisons, output_path, engine):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side as XlSide
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, PieChart, LineChart, Reference

    wb = Workbook()
    generated_at = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    # ── Styles ──
    all_border = Border(
        left=XlSide(style="thin"), right=XlSide(style="thin"),
        top=XlSide(style="thin"), bottom=XlSide(style="thin"),
    )
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    data_center = Alignment(horizontal="center", vertical="center")
    data_left = Alignment(horizontal="left", vertical="center")

    base_font = Font(name=FONT_NAME, size=FONT_SIZE)
    bold_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
    header_font = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")
    title_font = Font(name=FONT_NAME, size=16, bold=True)
    subtitle_font = Font(name=FONT_NAME, size=FONT_SIZE, italic=True, color="666666")
    section_font = Font(name=FONT_NAME, size=14, bold=True, color="2F5496")
    note_font = Font(name=FONT_NAME, size=10, italic=True, color="666666")

    header_fill = PatternFill("solid", fgColor="2F5496")
    match_fill = PatternFill("solid", fgColor="C6EFCE")
    mismatch_fill = PatternFill("solid", fgColor="FFC7CE")
    impact_fill = PatternFill("solid", fgColor="E2EFDA")

    inr_fmt = '₹#,##0.00'

    def apply_header(ws, row, start_col, count):
        for c in range(start_col, start_col + count):
            cell = ws.cell(row=row, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = all_border

    def dc(ws, row, col, value, fmt=None, align="center"):
        cell = ws.cell(row=row, column=col, value=value)
        cell.font = base_font
        cell.border = all_border
        cell.alignment = data_left if align == "left" else data_center
        if fmt:
            cell.number_format = fmt
        return cell

    def disable_gridlines(ws):
        ws.sheet_view.showGridLines = False

    def write_generated_at(ws, row, col):
        ws.cell(row=row, column=col,
                value=f"Report generated: {generated_at}").font = note_font

    broker = data.get("broker", "unknown")
    cdate = data.get("contract_date", "N/A")

    # ═══════════════════════════════════════════════════════════
    # SHEET 1: Contract Note Validation
    # ═══════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Contract Note Validation"
    ws1.sheet_properties.tabColor = "2F5496"
    disable_gridlines(ws1)

    row = SR
    ws1.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 7)
    ws1.cell(row=row, column=SC, value="NSE Cost Engine — Contract Note Validation Report").font = title_font
    row += 1
    ws1.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 7)
    ws1.cell(row=row, column=SC,
             value=f"Broker: {broker.title()} | Contract Date: {cdate}").font = subtitle_font
    row += 1
    write_generated_at(ws1, row, SC)

    # Trade details
    row += 2
    ws1.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 7)
    ws1.cell(row=row, column=SC, value="Trade Details").font = section_font
    row += 1
    t_headers = ["#", "Symbol", "Instrument", "Side", "Lot Size", "Qty (units)", "Premium (₹)", "Premium Value (₹)"]
    for c, h in enumerate(t_headers):
        ws1.cell(row=row, column=SC + c, value=h)
    apply_header(ws1, row, SC, len(t_headers))
    row += 1
    for i, (raw, trade, result) in enumerate(results, 1):
        vals = [i, raw["symbol"], raw.get("instrument", ""), raw["side"].upper(),
                trade.lot_size, trade.quantity * trade.lot_size,
                trade.premium if trade.premium else 0, trade.premium_value]
        for c, v in enumerate(vals):
            dc(ws1, row, SC + c, v, fmt=inr_fmt if c >= 6 else None)
        row += 1

    # Validation
    row += 1
    ws1.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 4)
    ws1.cell(row=row, column=SC, value="Component-wise Validation").font = section_font
    row += 1
    v_headers = ["Cost Component", "Engine (CN rounded)", "Contract Note", "Delta (₹)", "Status"]
    for c, h in enumerate(v_headers):
        ws1.cell(row=row, column=SC + c, value=h)
    apply_header(ws1, row, SC, len(v_headers))
    row += 1
    for comp in comparisons:
        dc(ws1, row, SC, comp["label"], align="left")
        dc(ws1, row, SC + 1, comp["engine"], fmt=inr_fmt)
        cn_val = comp["contract_note"]
        dc(ws1, row, SC + 2, cn_val if cn_val is not None else "N/A",
           fmt=inr_fmt if cn_val is not None else None)
        dc(ws1, row, SC + 3, comp["delta"] if comp["delta"] is not None else "",
           fmt='₹#,##0.00;(₹#,##0.00);"-"')
        cell = dc(ws1, row, SC + 4,
                  "✓ Match" if comp.get("match") else ("✗ Mismatch" if comp["match"] is not None else "No CN data"))
        if comp["match"] is not None:
            cell.fill = match_fill if comp["match"] else mismatch_fill
        row += 1

    total_eng = sum(c["engine"] for c in comparisons)
    total_cn = sum(c["contract_note"] for c in comparisons if c["contract_note"] is not None)
    dc(ws1, row, SC, "TOTAL", align="left").font = bold_font
    dc(ws1, row, SC + 1, total_eng, fmt=inr_fmt).font = bold_font
    dc(ws1, row, SC + 2, total_cn, fmt=inr_fmt).font = bold_font
    dc(ws1, row, SC + 3, total_eng - total_cn, fmt='₹#,##0.00;(₹#,##0.00);"-"')
    cell = dc(ws1, row, SC + 4, "✓ Match" if abs(total_eng - total_cn) < 0.10 else "✗ Mismatch")
    cell.fill = match_fill if abs(total_eng - total_cn) < 0.10 else mismatch_fill

    row += 2
    ws1.cell(row=row, column=SC,
             value="Rounding: STT & Stamp Duty → nearest ₹, all others → 2 decimals (source: Dhan)").font = note_font

    ws1.column_dimensions["A"].width = 3
    for c, w in enumerate([24, 20, 18, 14, 14, 16, 16, 18], 0):
        ws1.column_dimensions[get_column_letter(SC + c)].width = w

    # ═══════════════════════════════════════════════════════════
    # SHEET 2: Per-Trade Breakdown
    # ═══════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Per-Trade Breakdown")
    ws2.sheet_properties.tabColor = "548235"
    disable_gridlines(ws2)

    row = SR
    ws2.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 11)
    ws2.cell(row=row, column=SC, value="Cost Breakdown — Per Trade Leg").font = title_font
    row += 1
    write_generated_at(ws2, row, SC)

    row += 2
    b_headers = ["Symbol", "Side", "Premium Value", "Brokerage", "STT", "Exchange",
                  "SEBI", "Stamp Duty", "GST", "DP", "MTF Interest", "Total"]
    for c, h in enumerate(b_headers):
        ws2.cell(row=row, column=SC + c, value=h)
    apply_header(ws2, row, SC, len(b_headers))
    row += 1
    for raw, trade, result in results:
        d = result.as_dict()
        vals = [raw["symbol"], raw["side"].upper(), trade.premium_value,
                d["brokerage"], d["stt"], d["exchange_charges"], d["sebi_fee"],
                d["stamp_duty"], d["gst"], d["dp_charges"], d["mtf_interest"], d["total_cost"]]
        for c, v in enumerate(vals):
            dc(ws2, row, SC + c, v, fmt=inr_fmt if c >= 2 else None,
               align="left" if c <= 1 else "center")
        row += 1

    ws2.column_dimensions["A"].width = 3
    for c, w in enumerate([16, 8, 16, 12, 12, 12, 10, 12, 10, 10, 14, 14], 0):
        ws2.column_dimensions[get_column_letter(SC + c)].width = w

    # ═══════════════════════════════════════════════════════════
    # SHEET 3: Stress Tests
    # ═══════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Stress Tests")
    ws3.sheet_properties.tabColor = "BF8F00"
    disable_gridlines(ws3)

    row = SR
    ws3.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 10)
    ws3.cell(row=row, column=SC, value="Stress Test Scenarios").font = title_font
    row += 1
    write_generated_at(ws3, row, SC)

    scenarios = [
        ("Penny stock ₹2, 5000 shares intraday sell",
         dict(symbol="PENNY", segment=Segment.EQUITY, trade_type=TradeType.INTRADAY,
              side=TradeSide.SELL, exchange=Exchange.NSE, price=2.0, quantity=5000)),
        ("₹1Cr RELIANCE delivery buy",
         dict(symbol="RELIANCE", segment=Segment.EQUITY, trade_type=TradeType.DELIVERY,
              side=TradeSide.BUY, exchange=Exchange.NSE, price=2500.0, quantity=4000)),
        ("₹1Cr RELIANCE delivery sell",
         dict(symbol="RELIANCE", segment=Segment.EQUITY, trade_type=TradeType.DELIVERY,
              side=TradeSide.SELL, exchange=Exchange.NSE, price=2500.0, quantity=4000)),
        ("Nifty 1-lot futures sell",
         dict(symbol="NIFTY", segment=Segment.FUTURES, trade_type=TradeType.INTRADAY,
              side=TradeSide.SELL, exchange=Exchange.NSE, price=24500.0, quantity=1, lot_size=75)),
        ("Nifty 10-lot options sell ₹100 prem",
         dict(symbol="NIFTY", segment=Segment.OPTIONS, trade_type=TradeType.INTRADAY,
              side=TradeSide.SELL, exchange=Exchange.NSE, price=24500.0, quantity=10,
              lot_size=75, premium=100.0, strike_price=24400.0)),
        ("Deep OTM ₹1 premium BankNifty buy",
         dict(symbol="BANKNIFTY", segment=Segment.OPTIONS, trade_type=TradeType.INTRADAY,
              side=TradeSide.BUY, exchange=Exchange.NSE, price=52000.0, quantity=1,
              lot_size=30, premium=1.0, strike_price=55000.0)),
        ("MTF ₹25L, 30 days hold",
         dict(symbol="HDFCBANK", segment=Segment.EQUITY, trade_type=TradeType.MTF,
              side=TradeSide.BUY, exchange=Exchange.NSE, price=1600.0, quantity=1563,
              mtf_funding_pct=0.75, mtf_holding_days=30)),
        ("MTF ₹1Cr, 1 day (top slab)",
         dict(symbol="TCS", segment=Segment.EQUITY, trade_type=TradeType.MTF,
              side=TradeSide.BUY, exchange=Exchange.NSE, price=4000.0, quantity=2667,
              mtf_funding_pct=0.75, mtf_holding_days=1)),
    ]

    row += 2
    s_headers = ["Scenario", "Trade Value", "Brokerage", "STT", "Exchange", "SEBI",
                  "Stamp Duty", "GST", "DP", "MTF Interest", "Total Cost"]
    for c, h in enumerate(s_headers):
        ws3.cell(row=row, column=SC + c, value=h)
    apply_header(ws3, row, SC, len(s_headers))
    row += 1
    for name, kwargs in scenarios:
        trade = Trade(**kwargs)
        result = engine.calculate(trade)
        d = result.as_dict()
        vals = [name, trade.trade_value, d["brokerage"], d["stt"],
                d["exchange_charges"], d["sebi_fee"], d["stamp_duty"],
                d["gst"], d["dp_charges"], d["mtf_interest"], d["total_cost"]]
        for c, v in enumerate(vals):
            dc(ws3, row, SC + c, v, fmt=inr_fmt if c >= 1 else None,
               align="left" if c == 0 else "center")
        row += 1

    ws3.column_dimensions["A"].width = 3
    for c, w in enumerate([38, 16, 12, 12, 12, 10, 12, 10, 10, 14, 14], 0):
        ws3.column_dimensions[get_column_letter(SC + c)].width = w

    # ═══════════════════════════════════════════════════════════
    # FETCH LIVE MARKET DATA
    # ═══════════════════════════════════════════════════════════
    print("\nFetching live market data from Dhan API...")

    impact_stocks = {
        "RELIANCE": "2885",
        "HDFCBANK": "1333",
        "TCS": "11536",
        "INFY": "1594",
        "ICICIBANK": "4963",
        "SBIN": "3045",
    }

    live_data = fetch_live_impact_data(impact_stocks)

    # Fallback mock data if API unavailable
    if not live_data:
        print("  Using mock data for market impact sheets")
        live_data = {
            "RELIANCE": (0.0170, 20000000),
            "HDFCBANK": (0.0145, 15000000),
            "TCS": (0.0130, 5000000),
            "INFY": (0.0155, 12000000),
            "ICICIBANK": (0.0160, 18000000),
            "SBIN": (0.0180, 25000000),
        }

    # ═══════════════════════════════════════════════════════════
    # SHEET 4: Market Impact — Order Size Scenarios
    # ═══════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Impact vs Order Size")
    ws4.sheet_properties.tabColor = "C00000"
    disable_gridlines(ws4)

    # Use RELIANCE as the reference stock
    ref_symbol = "RELIANCE"
    ref_price = 2500.0
    ref_vol, ref_adv = live_data.get(ref_symbol, (0.017, 20000000))

    row = SR
    ws4.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 7)
    ws4.cell(row=row, column=SC,
             value=f"Market Impact Analysis — {ref_symbol} @ ₹{ref_price:,.0f}").font = title_font
    row += 1
    ws4.cell(row=row, column=SC,
             value=f"Daily Volatility: {ref_vol:.4f} | Avg Daily Volume: {ref_adv:,} shares").font = subtitle_font
    row += 1
    write_generated_at(ws4, row, SC)

    # Impact model explanation
    row += 2
    ws4.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 7)
    ws4.cell(row=row, column=SC,
             value="Square-Root Impact Model: Impact = σ × √(Q/V) × η  |  η = 0.3 (calibration constant)").font = note_font

    # Table: varying order sizes
    row += 2
    ws4.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 7)
    ws4.cell(row=row, column=SC, value="Impact at Different Order Sizes").font = section_font
    row += 1

    i_headers = ["Order Size (shares)", "Trade Value (₹)", "Participation Rate",
                  "Regulatory Costs (₹)", "Market Impact (₹)", "Total Cost (₹)",
                  "Impact as % of Regulatory", "Impact as bps of Trade"]
    for c, h in enumerate(i_headers):
        ws4.cell(row=row, column=SC + c, value=h)
    apply_header(ws4, row, SC, len(i_headers))
    data_start_row = row + 1
    row += 1

    order_sizes = [100, 500, 1000, 5000, 10000, 25000, 50000, 100000]

    for qty in order_sizes:
        trade = Trade(
            symbol=ref_symbol, segment=Segment.EQUITY,
            trade_type=TradeType.INTRADAY, side=TradeSide.BUY,
            exchange=Exchange.NSE, price=ref_price, quantity=qty,
            daily_volatility=ref_vol, avg_daily_volume=ref_adv,
        )
        result = engine.calculate(trade)
        participation = qty / ref_adv if ref_adv > 0 else 0
        reg_costs = result.total_cost_without_impact
        impact = result.market_impact
        impact_pct_reg = (impact / reg_costs * 100) if reg_costs > 0 else 0
        impact_bps = (impact / trade.trade_value * 10000) if trade.trade_value > 0 else 0

        vals = [qty, trade.trade_value, participation, reg_costs, impact,
                result.total_cost, impact_pct_reg, impact_bps]
        fmts = ["#,##0", inr_fmt, "0.00%", inr_fmt, inr_fmt, inr_fmt, "0.0\"%\"", "0.00\" bps\""]
        for c, (v, f) in enumerate(zip(vals, fmts)):
            dc(ws4, row, SC + c, v, fmt=f)
        row += 1

    data_end_row = row - 1

    # ── Chart 1: Impact vs Order Size (Bar) ──
    row += 1
    chart1 = BarChart()
    chart1.type = "col"
    chart1.style = 10
    chart1.title = f"Market Impact vs Regulatory Costs — {ref_symbol}"
    chart1.y_axis.title = "Cost (₹)"
    chart1.x_axis.title = "Order Size (shares)"
    chart1.width = 22
    chart1.height = 14

    cats = Reference(ws4, min_col=SC, min_row=data_start_row, max_row=data_end_row)
    reg_data = Reference(ws4, min_col=SC + 3, min_row=data_start_row - 1, max_row=data_end_row)
    imp_data = Reference(ws4, min_col=SC + 4, min_row=data_start_row - 1, max_row=data_end_row)

    chart1.add_data(reg_data, titles_from_data=True)
    chart1.add_data(imp_data, titles_from_data=True)
    chart1.set_categories(cats)

    chart1.series[0].graphicalProperties.solidFill = "2F5496"
    chart1.series[1].graphicalProperties.solidFill = "C00000"

    ws4.add_chart(chart1, f"{get_column_letter(SC)}{row}")

    # ── Chart 2: Impact as % of regulatory (Line) ──
    row += 18
    chart2 = LineChart()
    chart2.style = 10
    chart2.title = f"Market Impact as % of Regulatory Costs — {ref_symbol}"
    chart2.y_axis.title = "Impact / Regulatory (%)"
    chart2.x_axis.title = "Order Size (shares)"
    chart2.width = 22
    chart2.height = 14

    pct_data = Reference(ws4, min_col=SC + 6, min_row=data_start_row - 1, max_row=data_end_row)
    chart2.add_data(pct_data, titles_from_data=True)
    chart2.set_categories(cats)
    chart2.series[0].graphicalProperties.line.solidFill = "C00000"

    ws4.add_chart(chart2, f"{get_column_letter(SC)}{row}")

    # ── Chart 3: Pie chart for 10K order ──
    row += 18

    # Build pie data inline
    ws4.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 3)
    ws4.cell(row=row, column=SC,
             value=f"Cost Composition — 10,000 shares {ref_symbol}").font = section_font
    row += 1

    trade_10k = Trade(
        symbol=ref_symbol, segment=Segment.EQUITY,
        trade_type=TradeType.INTRADAY, side=TradeSide.BUY,
        exchange=Exchange.NSE, price=ref_price, quantity=10000,
        daily_volatility=ref_vol, avg_daily_volume=ref_adv,
    )
    r_10k = engine.calculate(trade_10k)

    pie_labels = ["Brokerage", "STT", "Exchange + SEBI + IPFT",
                   "Stamp Duty", "GST", "Market Impact"]
    pie_values = [
        r_10k.brokerage, r_10k.stt,
        r_10k.exchange_charges + r_10k.sebi_fee + r_10k.ipft,
        r_10k.stamp_duty, r_10k.gst, r_10k.market_impact,
    ]

    ws4.cell(row=row, column=SC, value="Component").font = bold_font
    ws4.cell(row=row, column=SC, value="Component").border = all_border
    ws4.cell(row=row, column=SC + 1, value="Amount (₹)").font = bold_font
    ws4.cell(row=row, column=SC + 1, value="Amount (₹)").border = all_border
    pie_data_start = row + 1
    row += 1

    for label, val in zip(pie_labels, pie_values):
        dc(ws4, row, SC, label, align="left")
        dc(ws4, row, SC + 1, round(val, 2), fmt=inr_fmt)
        row += 1
    pie_data_end = row - 1

    pie_chart = PieChart()
    pie_chart.title = f"Cost Breakdown — 10K shares {ref_symbol}"
    pie_chart.width = 16
    pie_chart.height = 12

    pie_cats = Reference(ws4, min_col=SC, min_row=pie_data_start, max_row=pie_data_end)
    pie_vals = Reference(ws4, min_col=SC + 1, min_row=pie_data_start, max_row=pie_data_end)
    pie_chart.add_data(pie_vals, titles_from_data=False)
    pie_chart.set_categories(pie_cats)

    ws4.add_chart(pie_chart, f"{get_column_letter(SC + 3)}{pie_data_start - 1}")

    ws4.column_dimensions["A"].width = 3
    for c, w in enumerate([22, 18, 18, 20, 18, 18, 22, 20], 0):
        ws4.column_dimensions[get_column_letter(SC + c)].width = w

    # ═══════════════════════════════════════════════════════════
    # SHEET 5: Cross-Stock Impact Comparison
    # ═══════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Impact Cross-Stock")
    ws5.sheet_properties.tabColor = "7030A0"
    disable_gridlines(ws5)

    row = SR
    ws5.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 8)
    ws5.cell(row=row, column=SC,
             value="Market Impact Comparison — 10,000 Shares Across Stocks").font = title_font
    row += 1
    ws5.cell(row=row, column=SC,
             value="Same order size (10,000 shares), different stocks — shows how liquidity affects impact").font = subtitle_font
    row += 1
    write_generated_at(ws5, row, SC)

    row += 2
    # Stock prices (approximate)
    stock_prices = {
        "RELIANCE": 2500, "HDFCBANK": 1800, "TCS": 3800,
        "INFY": 1500, "ICICIBANK": 1300, "SBIN": 800,
    }

    cs_headers = ["Stock", "Price (₹)", "Daily Vol", "Avg Daily Volume",
                   "Participation Rate", "Regulatory Costs (₹)", "Market Impact (₹)",
                   "Total Cost (₹)", "Impact as % of Regulatory"]
    for c, h in enumerate(cs_headers):
        ws5.cell(row=row, column=SC + c, value=h)
    apply_header(ws5, row, SC, len(cs_headers))
    cs_data_start = row + 1
    row += 1

    for symbol in ["RELIANCE", "HDFCBANK", "TCS", "INFY", "ICICIBANK", "SBIN"]:
        if symbol not in live_data:
            continue
        vol, adv = live_data[symbol]
        price = stock_prices.get(symbol, 1000)
        qty = 10000

        trade = Trade(
            symbol=symbol, segment=Segment.EQUITY,
            trade_type=TradeType.INTRADAY, side=TradeSide.BUY,
            exchange=Exchange.NSE, price=float(price), quantity=qty,
            daily_volatility=vol, avg_daily_volume=adv,
        )
        result = engine.calculate(trade)
        participation = qty / adv if adv > 0 else 0
        reg = result.total_cost_without_impact
        impact = result.market_impact
        impact_pct = (impact / reg * 100) if reg > 0 else 0

        vals = [symbol, price, vol, adv, participation, reg, impact,
                result.total_cost, impact_pct]
        fmts = [None, inr_fmt, "0.0000", "#,##0", "0.00%",
                inr_fmt, inr_fmt, inr_fmt, "0.0\"%\""]
        for c, (v, f) in enumerate(zip(vals, fmts)):
            dc(ws5, row, SC + c, v, fmt=f, align="left" if c == 0 else "center")
        row += 1

    cs_data_end = row - 1

    # ── Bar chart: Regulatory vs Impact per stock ──
    row += 1
    chart3 = BarChart()
    chart3.type = "col"
    chart3.style = 10
    chart3.title = "Regulatory Costs vs Market Impact — 10K Shares"
    chart3.y_axis.title = "Cost (₹)"
    chart3.width = 22
    chart3.height = 14

    cs_cats = Reference(ws5, min_col=SC, min_row=cs_data_start, max_row=cs_data_end)
    cs_reg = Reference(ws5, min_col=SC + 5, min_row=cs_data_start - 1, max_row=cs_data_end)
    cs_imp = Reference(ws5, min_col=SC + 6, min_row=cs_data_start - 1, max_row=cs_data_end)

    chart3.add_data(cs_reg, titles_from_data=True)
    chart3.add_data(cs_imp, titles_from_data=True)
    chart3.set_categories(cs_cats)

    chart3.series[0].graphicalProperties.solidFill = "2F5496"
    chart3.series[1].graphicalProperties.solidFill = "C00000"

    ws5.add_chart(chart3, f"{get_column_letter(SC)}{row}")

    # ── Bar chart: Impact as % of regulatory ──
    row += 18
    chart4 = BarChart()
    chart4.type = "col"
    chart4.style = 10
    chart4.title = "Market Impact as % of Regulatory — 10K Shares"
    chart4.y_axis.title = "Impact / Regulatory (%)"
    chart4.width = 22
    chart4.height = 14

    cs_pct = Reference(ws5, min_col=SC + 8, min_row=cs_data_start - 1, max_row=cs_data_end)
    chart4.add_data(cs_pct, titles_from_data=True)
    chart4.set_categories(cs_cats)
    chart4.series[0].graphicalProperties.solidFill = "7030A0"

    ws5.add_chart(chart4, f"{get_column_letter(SC)}{row}")

    ws5.column_dimensions["A"].width = 3
    for c, w in enumerate([14, 12, 12, 18, 18, 20, 18, 18, 22], 0):
        ws5.column_dimensions[get_column_letter(SC + c)].width = w

    # ═══════════════════════════════════════════════════════════
    # SHEET 6: Rate Schedule
    # ═══════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Rate Schedule")
    ws6.sheet_properties.tabColor = "548235"
    disable_gridlines(ws6)

    row = SR
    ws6.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 4)
    ws6.cell(row=row, column=SC,
             value=f"Rate Schedule — {broker.title()}, effective Oct 2024").font = title_font
    row += 1
    write_generated_at(ws6, row, SC)

    sections = [
        ("STT — Securities Transaction Tax", [
            ["Equity Delivery", "0.1%", "Trade Value", "Both", "Finance Act 2004, amended 2024"],
            ["Equity Intraday", "0.025%", "Trade Value", "Sell", ""],
            ["Futures", "0.05%", "Trade Value (Notional)", "Sell", "Revised Oct 2024"],
            ["Options", "0.15%", "Premium Value", "Sell", "Revised Oct 2024"],
            ["Options Exercise", "0.15%", "Intrinsic Value", "Exercise", ""],
            ["ETF Delivery", "0.001%", "Trade Value", "Sell", ""],
        ]),
        ("Exchange Transaction Charges (NSE)", [
            ["Equity", "0.0030699%", "Trade Value", "Both", ""],
            ["Futures", "0.0018299%", "Trade Value", "Both", ""],
            ["Options", "0.0355299%", "Premium Value", "Both", ""],
        ]),
        ("Stamp Duty (Buy Side Only)", [
            ["Equity Delivery", "0.015%", "Trade Value", "Buy", ""],
            ["Equity Intraday", "0.003%", "Trade Value", "Buy", ""],
            ["Futures", "0.002%", "Trade Value", "Buy", ""],
            ["Options", "0.003%", "Premium Value", "Buy", ""],
        ]),
        ("Other Charges", [
            ["SEBI Turnover Fee", "0.0001%", "Turnover", "Both", "₹10 per crore"],
            ["IPFT", "0.0000001%", "Turnover", "Both", "₹0.01 per crore"],
            ["GST", "18%", "Brokerage + Exch + SEBI + IPFT", "Both", ""],
            ["DP Charges (Dhan)", "₹12.50 flat", "Per ISIN per day", "Delivery Sell", "+18% GST"],
        ]),
        ("Market Impact (Square-Root Model)", [
            ["Formula", "σ × √(Q/V) × η", "η = 0.3", "Both", "Industry standard"],
            ["σ", "Daily volatility", "Std dev of log returns", "", "From Dhan API"],
            ["Q", "Order quantity", "Shares in the order", "", ""],
            ["V", "Avg daily volume", "30-day mean volume", "", "From Dhan API"],
        ]),
    ]

    row += 2
    for section_name, items in sections:
        ws6.merge_cells(start_row=row, start_column=SC, end_row=row, end_column=SC + 4)
        ws6.cell(row=row, column=SC, value=section_name).font = section_font
        row += 1
        r_headers = ["Segment", "Rate", "Base", "Side", "Notes"]
        for c, h in enumerate(r_headers):
            ws6.cell(row=row, column=SC + c, value=h)
        apply_header(ws6, row, SC, len(r_headers))
        row += 1
        for item in items:
            for c, v in enumerate(item):
                dc(ws6, row, SC + c, v, align="left" if c in (0, 4) else "center")
            row += 1
        row += 1

    ws6.column_dimensions["A"].width = 3
    for c, w in enumerate([22, 18, 30, 14, 24], 0):
        ws6.column_dimensions[get_column_letter(SC + c)].width = w

    wb.save(output_path)
    print(f"\n✓ Report saved: {output_path}")
    print(f"  6 sheets | 4 charts | generated at {generated_at}")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Generate NSE Cost Engine validation report")
    parser.add_argument("--input", "-i", help="Contract note JSON file", default=None)
    parser.add_argument("--output", "-o", help="Output Excel path", default=None)
    parser.add_argument("--broker", "-b", help="Broker profile", default=None)
    args = parser.parse_args()

    if args.input:
        input_path = Path(args.input)
        if not input_path.exists():
            print(f"Error: {input_path} not found")
            sys.exit(1)
        with open(input_path) as f:
            data = json.load(f)
    else:
        sample = Path(__file__).resolve().parent.parent / "tests" / "fixtures" / "contract_note_20260508.json"
        if sample.exists():
            with open(sample) as f:
                data = json.load(f)
        else:
            print("No input file specified and sample not found.")
            sys.exit(1)

    broker = args.broker or data.get("broker", "dhan")
    engine = CostEngine(broker=broker)

    trade_pairs = build_trades_from_json(data["trades"])
    actuals = data.get("contract_note_actuals", {})

    results, comparisons = run_validation(engine, trade_pairs, actuals)
    print_console_report(data, results, comparisons)

    if args.output:
        output_path = Path(args.output)
    else:
        output_path = Path(__file__).resolve().parent / "validation_report.xlsx"

    output_path.parent.mkdir(parents=True, exist_ok=True)

    if _check_openpyxl():
        generate_excel_report(data, trade_pairs, results, comparisons, str(output_path), engine)
    else:
        print("Install openpyxl for Excel: pip install openpyxl")


if __name__ == "__main__":
    main()